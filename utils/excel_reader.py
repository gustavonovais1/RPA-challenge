from __future__ import annotations
import logging
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

logger = logging.getLogger("rpa_challenge")


class ExcelReader:
    """Leitura de arquivos Excel."""

    def __init__(self, file_path: Path) -> None:
        self._file_path = Path(file_path)

    def path_exists(self) -> bool:
        return self._file_path.is_file()

    def read_rows(self) -> List[Dict[str, Any]]:
        """Lê as linhas da planilha como dicionários."""
        if not self.path_exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self._file_path}")

        logger.info("Lendo planilha: %s", self._file_path)

        wb = load_workbook(self._file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            iterator = ws.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if not header_row:
                logger.warning("Planilha sem linhas.")
                return []

            headers: List[str] = []
            for cell in header_row:
                if cell is None or str(cell).strip() == "":
                    headers.append("")
                else:
                    headers.append(str(cell).strip())

            if not any(headers):
                raise ValueError("Cabeçalho da planilha está vazio ou inválido.")

            rows_out: List[Dict[str, Any]] = []
            for row in iterator:
                if row is None:
                    continue
                if all(v is None or str(v).strip() == "" for v in row):
                    continue

                record: Dict[str, Any] = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    value = row[idx] if idx < len(row) else None
                    record[header] = value

                rows_out.append(record)

            logger.info("Total de linhas de dados lidas: %d", len(rows_out))
            return rows_out
        finally:
            wb.close()
